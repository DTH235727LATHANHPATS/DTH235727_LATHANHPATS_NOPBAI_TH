import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl

# ======= HÀM LƯU FILE EXCEL =======
def save_to_excel(data, file_path):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "NhanVien"
    sheet.append(["Mã", "Tên", "Tuổi"])
    for nv in data:
        sheet.append(nv)
    wb.save(file_path)

# ======= HÀM ĐỌC FILE EXCEL =======
def read_from_excel(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(list(row))
    return data

# ======= GIAO DIỆN CHÍNH =======
class EmployeeApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Quản lý Nhân viên (Excel)")
        self.root.geometry("600x400")

        self.data = []  # Danh sách nhân viên [ [Mã, Tên, Tuổi], ... ]
        self.file_path = "nhanvien.xlsx"

        # ====== Frame nhập liệu ======
        frame_input = tk.Frame(root)
        frame_input.pack(pady=10)

        tk.Label(frame_input, text="Mã:").grid(row=0, column=0, padx=5, pady=5)
        tk.Label(frame_input, text="Tên:").grid(row=0, column=2, padx=5, pady=5)
        tk.Label(frame_input, text="Tuổi:").grid(row=0, column=4, padx=5, pady=5)

        self.ma_entry = tk.Entry(frame_input, width=10)
        self.ten_entry = tk.Entry(frame_input, width=20)
        self.tuoi_entry = tk.Entry(frame_input, width=10)
        self.ma_entry.grid(row=0, column=1)
        self.ten_entry.grid(row=0, column=3)
        self.tuoi_entry.grid(row=0, column=5)

        tk.Button(frame_input, text="Thêm", command=self.add_employee).grid(row=0, column=6, padx=10)

        # ====== Treeview hiển thị ======
        columns = ("Mã", "Tên", "Tuổi")
        self.tree = ttk.Treeview(root, columns=columns, show="headings")
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=150, anchor="center")
        self.tree.pack(pady=10, fill="x")

        # ====== Các nút chức năng ======
        frame_btn = tk.Frame(root)
        frame_btn.pack(pady=10)

        tk.Button(frame_btn, text="Lưu Excel", command=self.save_file).grid(row=0, column=0, padx=10)
        tk.Button(frame_btn, text="Đọc Excel", command=self.load_file).grid(row=0, column=1, padx=10)
        tk.Button(frame_btn, text="Sắp xếp theo tuổi", command=self.sort_by_age).grid(row=0, column=2, padx=10)

    # ====== HÀM THÊM NHÂN VIÊN ======
    def add_employee(self):
        ma = self.ma_entry.get().strip()
        ten = self.ten_entry.get().strip()
        tuoi = self.tuoi_entry.get().strip()

        if not ma or not ten or not tuoi:
            messagebox.showwarning("Lỗi", "Vui lòng nhập đầy đủ thông tin.")
            return
        if not tuoi.isdigit():
            messagebox.showwarning("Lỗi", "Tuổi phải là số.")
            return

        self.data.append([ma, ten, int(tuoi)])
        self.tree.insert("", "end", values=(ma, ten, tuoi))

        self.ma_entry.delete(0, tk.END)
        self.ten_entry.delete(0, tk.END)
        self.tuoi_entry.delete(0, tk.END)
